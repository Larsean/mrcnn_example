# -*- coding: utf-8 -*-

import os
import random
from math import ceil
import re
import time
import numpy as np
import cv2
import matplotlib.pyplot as plt

import tensorflow as tf
from mrcnn.config import Config
from mrcnn import model as modellib, utils
from mrcnn import visualize
import yaml

from PIL import Image
import warnings
import argparse
from datetime import datetime
import pandas as pd
import shutil
from find_minimal_loss import read_logs
import openpyxl 
import os.path as osp
import json
from glob import glob
from imgaug import augmenters as iaa




warnings.filterwarnings("ignore")
os.environ["CUDA_VISIBLE_DEVICES"] = "0" # 原本是1但感覺很慢我改成0






class ShapesConfig(Config):
    """Configuration for training on the toy shapes dataset.
    Derives from the base Config class and overrides values specific
    to the toy shapes dataset.
    """
    # Give the configuration a recognizable name
    NAME = "shapes"

    # Train on 1 GPU and 8 images per GPU. We can put multiple images on each
    # GPU because the images are small. Batch size is 8 (GPUs * images/GPU).
    GPU_COUNT = 2
    IMAGES_PER_GPU = 8 # 水保的程式IMAGES_PER_GPU我都設1

    # Number of classes (including background)
    NUM_CLASSES = 1 + 4  # background + 1 shapes #如果你有超過一個label, 記得改1 + 2或者以上

    # Use small images for faster training. Set the limits of the small side
    # the large side, and that determines the image shape.
    IMAGE_MIN_DIM = 320
    IMAGE_MAX_DIM = 384

    # Use smaller anchors because our image and objects are small
    RPN_ANCHOR_SCALES = (8 * 6, 16 * 6, 32 * 6, 64 * 6, 128 * 6)  # anchor side in pixels

    # Reduce training ROIs per image because the images are small and have
    # few objects. Aim to allow ROI sampling to pick 33% positive ROIs.
    TRAIN_ROIS_PER_IMAGE = 10

    # Use a small epoch since the data is simple
    STEPS_PER_EPOCH = 100

    # use small validation steps since the epoch is small
    VALIDATION_STEPS = 21

    # BACKBONE = "resnet50"





class DrugDataset(utils.Dataset):
    # 得到该图中有多少个实例（物体）
    def get_obj_index(self, image):
        n = np.max(image)
        return n

    # 解析labelme中得到的yaml文件，从而得到mask每一层对应的实例标签
    def from_yaml_get_class(self, image_id):
        info = self.image_info[image_id]
        with open(info['yaml_path']) as f:
            temp = yaml.load(f.read(), Loader=yaml.FullLoader)
            labels = temp['label_names']
            del labels[0]
        return labels

    # 重新写draw_mask
    def draw_mask(self, num_obj, mask, image, image_id, ):

        info = self.image_info[image_id]
        img = np.array(image)
        mask = (np.arange(img.max()) == img[..., None] - 1).astype(int)

        # for index in range(num_obj):
        #     for i in range(info['width']):
        #         for j in range(info['height']):
        #             at_pixel = image.getpixel((i, j))
        #             if at_pixel == index + 1:
        #                 mask[j, i, index] = 1

        return mask



    # 重新写load_shapes，里面包含自己的自己的类别
    # 并在self.image_info信息中添加了path、mask_path 、yaml_path
    # yaml_pathdataset_root_path = "/tongue_dateset/"
    # img_floder = dataset_root_path + "rgb"
    # mask_floder = dataset_root_path + "mask"
    # dataset_root_path = "/tongue_dateset/"
    # 原本的 for i in range(count)
    # 現在是 for i in range(start, end)
    def load_shapes(self, img_floder, mask_floder, dataset_root_path, class_names, imglist):
        """Generate the requested number of synthetic images.
        count: number of images to generate.
        height, width: the size of the generated images.
        """
        
        self.new_class = {label:i+1 for i, label in enumerate(class_names[1:])}
        count = len(imglist)


        for Class_name, Index in self.new_class.items():
            self.add_class("shapes", Index, Class_name)



        for i in range(count):
            # 获取图片宽和高
            print(i)
            filestr = imglist[i].split(".")[0]
            mask_path = mask_floder + "/" + filestr + ".png"
            yaml_path = dataset_root_path + "/labelme_json/" + filestr + "_json/info.yaml"
            meta_path = dataset_root_path + "/meta/" + filestr + ".npz"
            print(dataset_root_path + "/labelme_json/" + filestr + "_json/img.png")
            cv_img = cv2.imread(dataset_root_path + "/labelme_json/" + filestr + "_json/img.png")

            self.add_image("shapes", image_id=i, path=img_floder + "/" + imglist[i],
                           width=cv_img.shape[1], height=cv_img.shape[0], mask_path=mask_path, yaml_path=yaml_path, meta_path=meta_path)

    # 重写load_mask
    def load_mask(self, image_id):
        """Generate instance masks for shapes of the given image ID.
        """
        print("image_id", image_id)
        info = self.image_info[image_id]

        if osp.exists(info["meta_path"]):
            with np.load(info["meta_path"]) as data:
                data['']
                class_ids, mask = data["class_ids"], data["mask"]
            return mask, class_ids
        img = Image.open(info['mask_path'])
        num_obj = self.get_obj_index(img)
        count = num_obj # 原本是count=1
        mask = np.zeros([info['height'], info['width'], num_obj], dtype=np.uint8)

        mask = self.draw_mask(num_obj, mask, img, image_id)
        occlusion = np.logical_not(mask[:, :, -1]).astype(np.uint8)
        for i in range(count - 2, -1, -1):
            mask[:, :, i] = mask[:, :, i] * occlusion

            occlusion = np.logical_and(occlusion, np.logical_not(mask[:, :, i]))

        labels = self.from_yaml_get_class(image_id)
        class_ids = np.array([self.new_class[label_name] for label_name in labels])
        np.savez(info["meta_path"], mask=mask, class_ids=class_ids)
        return mask, class_ids.astype(np.int32)

def kfold(data_list, ratio):
    data_list = np.array(data_list) 
    _, val_ratio, test_ratio = ratio
    length = len(data_list)
    total = sum(ratio)
    
    data_kfold = []
    for i in range(total):
        begin = ceil(length*i/total)
        val_end = begin + ceil(length*val_ratio/total)
        test_end = val_end + ceil(length*test_ratio/total)
        val_ind = np.array([i if i<length else i-length  for i in np.arange(begin,val_end)])
        test_ind = np.array([i if i<length else i-length  for i in np.arange(val_end,test_end)])
        train_ind = np.setdiff1d(np.arange(length), np.concatenate([val_ind, test_ind]))
        data_kfold.append((data_list[train_ind].tolist(), data_list[val_ind].tolist(), data_list[test_ind].tolist()))
    return data_kfold


def monte_carlo(data_list, ratio, times):
    _, val_ratio, test_ratio = ratio
    val_amount = ceil(len(data_list) * val_ratio / sum(ratio))
    test_amount = ceil(len(data_list) * test_ratio / sum(ratio))

    data_monte = []
    for i in range(times):
        # 先將val,test全部抽出，剩餘為train，val在從暫存抽出所需
        temp = random.sample(data_list, val_amount + test_amount)
        data_val = random.sample(temp, val_amount)
        data_test = list(set(temp) - set(data_val))
        data_train = list(set(data_list) - set(temp))
        data_monte.append((data_train, data_val, data_test))
    return data_monte


def Find_last_weight(weights_folder):
    cur_epoch = [0,"weights_name(.h5)"]
    for file in os.listdir(weights_folder):
        if ".h5" in file:
            file_epoch = re.search("(\d{4}).h5", file).group(1)
            file_epoch = int(file_epoch)
            cur_epoch = [file_epoch, file] if cur_epoch[0]<file_epoch else cur_epoch

    assert cur_epoch[0]!=0 , "Not found Model_weights.h5" 

    return  osp.join(weights_folder, cur_epoch[1])

def extract_Info(Dict):
    save_property = ["dir_path", "class_names", "sampling_mode", "sampling_times", "save_path", "ratio"
                    ,"special_sample_list"]
    info = {key:Dict[key] for key in save_property}
    info["root_path"] = info.pop("dir_path")
    return info

def get_weights_path(init_with, Model_dir, info):
    weights_path_list, save_weights_folders = [], []
    
    if init_with == "resume":
        logs_path = osp.join(Model_dir, "Temp", "loss.xlsx")
        book = pd.read_excel(logs_path,sheet_name=None ,engine="openpyxl")
        if (not info["special_sample_list"]) and info['sampling_mode']!="kfold":
            raise ValueError('special_sample_list is empty')
        if info["special_sample_list"]:
            sample_sheets = info["special_sample_list"]
        elif info['sampling_mode'] == "kfold":
            num_partition = sum(info["ratio"]) 
            sample_sheets = [f"kfold_{times+1}" for times in range(num_partition)]
        for sheet in sample_sheets:
            if sheet in book:
                df = book[sheet]
                sheet_info = df.loc[0, 'info']
                sheet_info = eval(sheet_info)
                #修正
                # weights_path = Find_last_weight(sheet_info['save_path'])
                # save_folder = osp.join(sheet_info['save_path'])
                #儲存位置有更動則執行下方程式
                weights_path = Find_last_weight(f"{info['save_path']}/{sheet}")
                save_folder = osp.join(info['save_path'], sheet)
            elif info['sampling_mode'] == "kfold":
                weights_folder = osp.join(info['save_path'], "inital_weights")
                weights_path = glob(f"{weights_folder}/*.h5")[0]
                save_folder = osp.join(info['save_path'], sheet)
            weights_path_list.append(weights_path)
            save_weights_folders.append(save_folder)
        return weights_path_list, save_weights_folders
    
    inital_folder = osp.join(info['save_path'], "inital_weights")
    if init_with == "coco":
        weights_path = osp.join("mask_rcnn_coco.h5")
    elif init_with == "last":
        for folder in os.listdir(Model_dir):
            regex = r"[\w-]+(\d{4})(\d{2})(\d{2})T(\d{2})(\d{2})"
            m = re.match(regex, folder)
            if not m:
                continue
            folder_date = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                                int(m.group(4)), int(m.group(5)))


            if ('record_date' not in dir()) or folder_date > record_date :
                record_date = folder_date
                last_folder = osp.join(Model_dir, folder)
        weights_path = Find_last_weight(last_folder)
    elif init_with == "random_init":
        weights_list = glob(f"{Model_dir}/*/*.h5")
        weights_list.extend(glob(f"{info['save_path']}/*/*.h5"))
        weights_path = random.choice(weights_list)
    elif init_with=="sampling_init":
        weights_path = osp.join(inital_folder, os.listdir(inital_folder)[-1])

    if not osp.exists(inital_folder):
        weights_name = osp.split(weights_path)[-1]
        m = re.search("(\d{4}).h5", weights_name)
        if m:
            epoch = m.group(1)
            weights_name = weights_name.replace(epoch, "0000")
        inital_weights_path = osp.join(inital_folder, weights_name)
        os.makedirs(inital_folder)
        shutil.copyfile(weights_path, inital_weights_path)


    if info["sampling_mode"]=="kfold":
        sampling_times, save_folder = sum(info["ratio"]), f"{info['sampling_mode']}"
    elif info["sampling_mode"]=="monte_carlo" :
        sampling_times, save_folder = info["sampling_times"], f"{ info['date'] }_{ info['sampling_mode'] }"
    
    for times in range(sampling_times):
        save_path = osp.join(info['save_path'], f"{save_folder}_{times+1}")
        weights_path_list.append(weights_path)
        save_weights_folders.append(save_path)
    return weights_path_list, save_weights_folders
        


    
def get_epoch(weights_path, sheets_path=""):
    if osp.split(weights_path)[-1]=="mask_rcnn_coco.h5":
        return 0
    elif osp.exists(sheets_path):
        weights_epoch = re.search("(\d{4}).h5", weights_path).group(1)
        regex = r"([\w\-]+)[/\\\.]"
        m = re.findall(regex, weights_path)
        book = openpyxl.load_workbook(sheets_path)
        sheet_name = m[-2]
        sheet = book[sheet_name]
        sheet_epoch = sheet[f'A{sheet.max_row}'].value
        return int(weights_epoch) if int(weights_epoch)>int(sheet_epoch) else int(sheet_epoch)
    else:
        weights_epoch = re.search("(\d{4}).h5", weights_path).group(1)
        return int(weights_epoch)

def get_data_sampling(init_with, root_path, info, dataset_path=""):
    if init_with=="resume" :
        dataset_path = osp.join(dataset_path, "dataset.json")
        if osp.exists(dataset_path):
            with open(dataset_path) as f:
                data_sampling = json.load(f)
            return data_sampling


        elif "kfold" in osp.split(dataset_path)[-1]:
            times = osp.split(dataset_path)[-1].split("_")[-1]
            data_list = kfold(os.listdir(root_path), info['ratio'])
            data_sampling = data_list[int(times) - 1]
            if not osp.exists(dataset_path):
                os.makedirs(dataset_path)
                with open(osp.join(dataset_path, "dataset.json"), "w") as f:
                    f.write(json.dumps(data_sampling, indent=2))
            return data_sampling
        else:
            raise IOError(dataset_path+"\nnot found")
    if info["sampling_mode"]=="kfold":
        data_list = kfold(os.listdir(root_path), info['ratio'])
        return data_list[info["times"]-1]
    if info["sampling_mode"]=="monte_carlo":
        data_list = monte_carlo(os.listdir(root_path), info['ratio'], 1)
        return data_list[0]
    
def model_load_weights(init_with, model, weights_path, temp_folder):
    weights_name = osp.split(weights_path)[-1]

    m = re.search("(\d{4}).h5", weights_name)
    if m:
        if init_with == "resume":
            temp_weights = osp.join(temp_folder, weights_name)
        else:
            epoch = m.group(1)
            weights_name = weights_name.replace(epoch, "0000")
            temp_weights = osp.join(temp_folder, weights_name)
        shutil.copyfile(weights_path, temp_weights)
        model.load_weights(temp_weights, by_name=True)
    # init_with=="coco" or "sampling_init"(initial weights is mask_rcnn_coco.h5)
    else:
        temp_weights = osp.join(temp_folder, weights_name)
        shutil.copyfile(weights_path, temp_weights)
        model.load_weights(temp_weights, by_name=True,
                           exclude=["mrcnn_class_logits", "mrcnn_bbox_fc", "mrcnn_bbox", "mrcnn_mask"])
    print(temp_weights)
    os.remove(temp_weights)


def save_weights_and_loss(info, save_folder, resume=False):
    logs_excel = "loss.xlsx" if info["sampling_mode"] == "kfold" else "monte_loss.xlsx"
    save_items = ['class_names', 'root_path', 'save_path', 'times']
    current_folder = info["current_folder"]
    sheet_name = osp.split(save_folder)[-1]
    info = info.copy()
    info['save_path'] = save_folder
    info = {item:info[item] for item in save_items}
    loss_table = read_logs(current_folder)
    loss_table["info"] = [str(info)] + [None]*(loss_table.shape[0]-1)

    if resume:
        writer = pd.ExcelWriter(osp.join(current_folder, logs_excel), engin='openpyxl')
        book = openpyxl.load_workbook(writer.path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        original_sheet = book[sheet_name]
        data = original_sheet.values
        cols = next(data)
        original_table = pd.DataFrame(data, columns=cols)
        original_table.set_index("epochs", inplace=True)

        del loss_table['info']
        loss_table = pd.concat([original_table, loss_table], axis=0)
        loss_table = loss_table[~loss_table.index.duplicated(keep='last')]
        loss_table.to_excel(writer, sheet_name)
        writer.save()
        writer.close()
    elif osp.exists(osp.join(current_folder, logs_excel)):
        writer = pd.ExcelWriter(osp.join(current_folder, logs_excel), engin='openpyxl')
        writer.book = openpyxl.load_workbook(writer.path)
        loss_table.to_excel(writer,sheet_name=sheet_name)
        writer.save()
        writer.close()
    else:
        loss_table.to_excel(osp.join(current_folder, logs_excel), sheet_name=sheet_name)

    for file in os.listdir(current_folder):
        file_path = osp.join(current_folder, file)
        save_path =  osp.join(save_folder, file)
        if ".xlsx" in file:
            continue
        shutil.move(file_path, save_path)
        print(save_path)
        #print(file_path, save_path, sep="\n")


if __name__ == '__main__' :
    parser = argparse.ArgumentParser(description='Process Config')
    parser.add_argument('--config', help="Path of config.yaml", default="config.yaml")
    arg = parser.parse_args()
    path = arg.config

    try:
        with open(r"{}".format(path)) as f:    
            inital = yaml.load(f.read(), Loader=yaml.FullLoader)
            class_names = inital['class_names']
            dataset_root_path = inital['dir_path']
            img_floder = osp.join( dataset_root_path, "pic")  
            mask_floder = osp.join( dataset_root_path, "cv2_mask")
            MODEL_DIR = inital['logs_path']
            init_with = inital['init_with']
            epochs_per_sampling = inital["epochs_per_sampling"]
            info = extract_Info(inital)

    except:    
        raise IOError("Please input valid path and file(config.yaml)")




# 基础设置  修改 餵進去的訓練資料位置
config = ShapesConfig()

config.NUM_CLASSES = len(class_names)
config.IMAGE_META_SIZE = 1 + 3 + 3 + 4 + 1 + len(class_names)


info["current_folder"] = osp.join(MODEL_DIR, "Temp")
info["date"] = datetime.now().strftime("%m-%d-%H%M")

weights_path_list, save_weights_folders = get_weights_path(init_with, MODEL_DIR, info)
load_shape_config = [img_floder, mask_floder, dataset_root_path, class_names]
end_epoch = 0; epoch=0

aug = iaa.Sometimes(8/9,iaa.SomeOf((1,3),[iaa.Fliplr(1),iaa.Flipud(1),
            iaa.Affine(rotate=(-45, 45)), iaa.Affine(rotate=(-90, 90)),
            iaa.Affine(scale=(0.5, 1.5)), iaa.Add(15), iaa.Sharpen(alpha=0.4),iaa.GaussianBlur(1.0),
            iaa.GammaContrast(0.5,1) ] ) )

for weights_path, save_folder in zip(weights_path_list, save_weights_folders):
    print(weights_path, save_folder)


for times,(weights_path, save_folder) in enumerate(zip(weights_path_list, save_weights_folders)) :

    start =time.time()
    info.update({"times": times + 1})
    if init_with == "resume":
        cur_epoch = get_epoch(weights_path, osp.join(info['current_folder'], "loss.xlsx"))
        data_sampling = get_data_sampling(init_with, img_floder, info, save_folder)
    else:
        cur_epoch = 0
        data_sampling = get_data_sampling(init_with, img_floder, info)
        if not osp.exists(save_folder):
            os.makedirs(save_folder)
            with open(osp.join(save_folder, "dataset.json"), "w") as f:
                f.write(json.dumps(data_sampling, indent=2))
    train_list, val_list, test_list = data_sampling
    print(len(train_list), len(val_list), len(test_list))
    end_epoch = int(np.maximum(end_epoch, cur_epoch + epochs_per_sampling))
    print(weights_path)
    print(save_folder)
    print(end_epoch)
    print(save_weights_folders)

    model = modellib.MaskRCNN(mode="training", config=config,
                              model_dir=MODEL_DIR)  # 會生成這個資料夾裝log檔
    model_load_weights(init_with, model, weights_path, info["current_folder"])




    dataset_train = DrugDataset()
    dataset_train.load_shapes(*load_shape_config, train_list)
    dataset_train.prepare()

    dataset_val = DrugDataset()
    dataset_val.load_shapes(*load_shape_config, val_list) 
    dataset_val.prepare()

    dataset_test = DrugDataset()
    dataset_test.load_shapes(*load_shape_config, test_list) 
    dataset_test.prepare()



    # Train the head branches
    # Passing layers="heads" freezes all layers except the head
    # layers. You can also pass a regular expression to select
    # which layers to train by name pattern.
    model.train(dataset_train, dataset_val,
                learning_rate=config.LEARNING_RATE,
                epochs=end_epoch,  # 改成你要幾個epoch
                layers='heads',
                augmentation= aug
                )

    # Fine tune all layers
    # Passing layers="all" trains all layers. You can also
    # pass a regular expression to select which layers to
    # train by name pattern.
    model.train(dataset_train, dataset_val,
                learning_rate=config.LEARNING_RATE / 10,
                epochs=end_epoch,
                layers="all",
                augmentation= aug)
                
    save_weights_and_loss(info, save_folder, init_with=="resume")
    print("cost :", time.time()-start)